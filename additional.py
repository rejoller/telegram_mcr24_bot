import re
from openpyxl import Workbook
from io import BytesIO
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils import get_column_letter

def normalize_text_v2(text):
    text = text.lower().replace('ё', 'е').replace('р-н', 'район').replace('-', ' ')
    text = re.sub(r'(N|№|No)', 'N', text, flags=re.IGNORECASE)
    text = text.replace(' район', '').strip()
    return text
               
              
                
def split_message(message, max_length=4096):
    if len(message) <= max_length:
        return [message]

    messages = []
    while len(message) > max_length:
        split_index = message[:max_length].rfind('\n')
        if split_index == -1:
            split_index = max_length

        messages.append(message[:split_index])
        message = message[split_index:].lstrip()

    if message:
        messages.append(message)

    return messages
    
     
    
def create_excel_file_2(data):
    # Добавляем нумерацию в первый столбец и сортируем по количеству голосов

    #sorted_data = sorted(data, key=lambda x: int(x[1]) if (len(x) > 1 and str(x[1]).isdigit()) else 0, reverse=True)

    sorted_data = sorted(data, key=lambda x: x[4] if len(x) > 4 else 0, reverse=True)


    # Отфильтровываем пустые строки
    sorted_data = [row for row in sorted_data if any(row)]

    # Добавляем нумерацию в первый столбец
    data = [[i+1] + row for i, row in enumerate(sorted_data)]

    wb = Workbook()
    ws = wb.active

    headers = ['Позиция', 'Наименование населенного пункта', 'Количество голосов', 'Время обновления', 'Комментарий Минцифры России']

    header_font = Font(name='Arial', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='Arial')
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # Изменяем цвет заливки заголовка на светло-голубой
    header_fill = PatternFill(start_color='add8e6', end_color='add8e6', fill_type='solid')

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 20
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
        cell.fill = header_fill

    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_data)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file