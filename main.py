#!/home/support/telegram_mcr24_bot/myenv/bin/python

import openai
import aiogram
import logging
import asyncio


from aiogram import Bot, Dispatcher, types
from aiogram.utils import executor
from aiogram.contrib.middlewares.logging import LoggingMiddleware
from google.oauth2 import service_account
from googleapiclient.discovery import build

import openpyxl
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils import get_column_letter
from io import BytesIO

import os
import matplotlib.pyplot as plt
import numpy as np

import pandas as pd
import gspread
import requests
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from html import escape
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google.oauth2 import service_account
import aiohttp
import gspread_asyncio

from apscheduler.schedulers.asyncio import AsyncIOScheduler
from aiogram.utils.executor import start_webhook
from aiogram.contrib.fsm_storage.redis import RedisStorage2
from aiogram.contrib.fsm_storage.memory import MemoryStorage
import markdown


import shutil
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command
from aiogram.dispatcher.filters.state import State, StatesGroup

from google.oauth2 import service_account
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
import folium
from aiogram.types import InputFile
from folium.plugins import MarkerCluster
import random
import threading

import requests
from config import bot_token, SPREADSHEET_ID

response_storage = {}
bot = Bot(token=bot_token)
dp = Dispatcher(bot, storage=MemoryStorage())
info_text_storage = {}









@dp.message_handler(commands=['help'])
async def handle_help_command(message: types.Message):
    await log_user_data_from_message(message)
    help_text = (
        'Введи название населенного пункта или муниципального образования, чтобы получить информацию о связи. Чтобы узнать информацию о сотовой связи, выбери /2g /3g или /4g. Чтобы получить информацию о населенных пунктах без сотовой связи жми /nomobile\n\n'
        'Для получения списка ФАП из контракта с ПАО "Ростелеком" нажми /fp\n'
        'Для получения списка точек Аг.ГОиЧС из контракта с ПАО "Ростелеком" нажми /ago\n\n'
        'Чтобы узнать о подключении к ТОРКНД, введи сообщение "тор" и наименование муниципального образования. '
        'Например, "тор Енисейский".\n'
        'Если нужна статистика по всему краю, жми /knd_kraj\n\n'
        'Чтобы узнать, кто сегодня в отпуске, жми /otpusk\n\n'
        'Если остались вопросы, пиши @rejoller.')
    await message.reply(help_text)





SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
SERVICE_ACCOUNT_FILE = 'credentials.json'

creds = service_account.Credentials.from_service_account_file(
    SERVICE_ACCOUNT_FILE, scopes=SCOPES)


headers = ['Наименование', 'Население', 'Сотовая связь', 'Интернет', 'Программа', 'Таксофон', 'СЗО (узел)']


from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook







def adjust_column_width(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            value = str(cell.value)
            length = len(value.encode('utf-8'))
            if length > max_length:
                max_length = length

        # Настройка ширины столбца
        estimated_width = max_length * 0.7  # Умножение на коэффициент для учета разных ширин символов
        worksheet.column_dimensions[column].width = estimated_width


def convert_to_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = False
    # Настраиваем стили для заголовков
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    header_border = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="AED6F1",
                              end_color="AED6F1",
                              fill_type="solid")

    # Настраиваем стили для данных
    data_font = Font(size=11)
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    data_fill = PatternFill(start_color="ECECEC",
                            end_color="ECECEC",
                            fill_type="solid")

    for row_idx, row in enumerate(data, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Применяем стили
            if row_idx == 1:
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = header_border
                cell.fill = header_fill
            else:
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = data_border
                if row_idx % 2 == 0:
                    cell.fill = data_fill

    # Вызов функции для автоматической настройки ширины столбцов
    adjust_column_width(ws)

    # Добавляем автофильтр
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Закрепляем строку заголовка
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer












async def filter_and_send_data(message, filter_func, command):
    goroda_data = load_goroda_data()
    headers = ['Наименование населенного пункта', 'Население 2010', 'Население 2020', 'Сотовая связь', 'Программа']
    filtered_data = [headers]
    filtered_goroda_data = []

    for row in goroda_data:
        if filter_func(row):
            filtered_row = [row[i] if i < len(row) else '' for i in [1, 2, 5, 3, 11]]
            filtered_data.append(filtered_row)
            filtered_goroda_data.append(row)

    # Создать карту с маркерами
    map_with_markers = await create_map_with_markers(filtered_goroda_data)
    map_filename = f"{command}_map.html"
    map_with_markers.save(map_filename)

    # Конвертировать данные в формат Excel и отправить
    buffer = convert_to_excel(filtered_data)
    filename = f"{command}.xlsx"
    with open(filename, "wb") as excel_file:
        excel_file.write(buffer.getvalue())

    with open(filename, "rb") as excel_file:
        document = InputFile(excel_file)
        await message.answer_document(document=document, caption="Список населенных пунктов")

    os.remove(filename)

    # Отправить файл с картой
    # with open(map_filename, "rb") as map_file:
    #    document = InputFile(map_file)
    #  bot.send_document(message.chat.id, document=document, caption=map_filename)

    os.remove(map_filename)
    url = f"https://rejoller.pythonanywhere.com/{command}"
    await message.answer("Чтобы посмотреть карту, нажмите кнопку ниже", reply_markup=webAppKeyboard(url))






















def create_pie_chart(yes_count, no_count, filename):
    labels = ['Есть', 'Нет']
    sizes = [yes_count, no_count]
    colors = ['#2ecc71', '#e74c3c']

    # Создайте объект figure с заданными размерами (ширина, высота) в дюймах
    plt.figure(figsize=(2, 2))  # Здесь 2.5 дюйма - это ширина и высота диаграммы

    plt.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
    plt.axis('equal')
    plt.savefig(filename, dpi=400, bbox_inches='tight')  # Установите разрешение (DPI) изображения и параметр bbox_inches
    plt.clf()


def create_bar_chart(data, filename):
    labels, yes_values, no_values = zip(*data)

    total_values = [yes + no for yes, no in zip(yes_values, no_values)]
    yes_percentages = [yes / total * 100 if total != 0 else 0 for yes, total in zip(yes_values, total_values)]
    no_percentages = [no / total * 100 if total != 0 else 0 for no, total in zip(no_values, total_values)]

    labels = labels[1:]
    yes_percentages = yes_percentages[1:]
    no_percentages = no_percentages[1:]

    y = np.arange(len(labels))
    width = 0.6
    colors = ['#2ecc71', '#e74c3c']

    fig, ax = plt.subplots(figsize=(12, 16), dpi=300)  # Устанавливаем размер и DPI изображения
    rects1 = ax.barh(y, yes_percentages, label='Процент подключенных услуг', color=colors[0], align='center')
    rects2 = ax.barh(y, no_percentages, label='Процент не подключенных услуг', left=yes_percentages, color=colors[1],
                     align='center')

    ax.set_title('Подключение к ТОРКНД в Красноярском крае')
    ax.set_yticks(y)
    ax.set_yticklabels(labels)
    ax.legend()

    xmin = 0
    xmax = 100
    ax.set_xlim([xmin, xmax])

    def autolabel(rects, labels):
        for rect, label in zip(rects, labels):
            width = rect.get_width()
            ax.annotate('{:.1f}%'.format(label),
                        xy=(width, rect.get_y() + rect.get_height() / 2),
                        xytext=(3, 0),
                        textcoords="offset points",
                        ha='left', va='center')

    autolabel(rects1, yes_percentages)
    # autolabel(rects2, no_percentages)

    plt.tight_layout()
    plt.savefig(filename)
    plt.close()


async def search_szofed_values(column_4_value, spreadsheet):
    result = await spreadsheet.values_batch_get('szofed!A1:M2412')
    rows = result.get('valueRanges', [])[0].get('values', [])
    found_values = [row for row in rows if column_4_value.lower() == row[0].lower()]
    return found_values


@dp.message_handler(commands=['otpusk'])
async def handle_otpusk_command(message: types.Message):
    await message.answer('Загружаю данные')
    otpusk_data = await load_otpusk_data()
    await filter_and_send_data(message, otpusk_data)


async def filter_and_send_data(message: types.Message, data, headers=None):
    if headers:
        filtered_data = [headers] + data
    else:
        filtered_data = data

    filtered_data = [list(map(str, row)) for row in filtered_data]

    # Отправляем данные частями, чтобы не превысить лимит сообщения
    for i in range(0, len(filtered_data), 10):
        chunk = filtered_data[i:i+10]
        text = '\n'.join(['\t'.join(row) for row in chunk])
        await message.answer(text, parse_mode=types.ParseMode.MARKDOWN)


@dp.message_handler(commands=['employees_vacation'])
async def handle_employees_vacation_command(message: types.Message):
    await message.answer('Загружаю данные')
    otpusk_data = await load_otpusk_data()
    employees_on_vacation, employees_starting_vacation_soon = get_employees_on_vacation(otpusk_data)

    if employees_on_vacation:
        await message.answer('Сотрудники, находящиеся в отпуске:')
        await message.answer('\n'.join(['\t'.join(row) for row in employees_on_vacation]), parse_mode=types.ParseMode.MARKDOWN)
    else:
        await message.answer('Сотрудников в отпуске нет.')

    if employees_starting_vacation_soon:
        await message.answer('Сотрудники, начинающие отпуск в ближайшие дни:')
        await message.answer('\n'.join(['\t'.join(row) for row in employees_starting_vacation_soon]), parse_mode=types.ParseMode.MARKDOWN)
    else:
        await message.answer('Сотрудников, начинающих отпуск в ближайшие дни, нет.')


@dp.message_handler(commands=['pie_chart'])
async def handle_pie_chart_command(message: types.Message):
    await message.answer('Создаю круговую диаграмму')
    create_pie_chart(20, 80, 'pie_chart.png')
    with open('pie_chart.png', 'rb') as photo:
        await message.answer_photo(photo)


@dp.message_handler(commands=['bar_chart'])
async def handle_bar_chart_command(message: types.Message):
    await message.answer('Создаю гистограмму')
    data = [
        ('Район 1', 5, 10),
        ('Район 2', 10, 15),
        ('Район 3', 20, 5),
        ('Район 4', 30, 25),
        ('Район 5', 50, 20),
    ]
    create_bar_chart(data, 'bar_chart.png')
    with open('bar_chart.png', 'rb') as photo:
        await message.answer_photo(photo)













#11


def escape_markdown(text):
    markdown_escape_characters = ['*', '_', '[', ']', '(', ')', '~', '`', '>', '#', '+', '-', '=', '|', '{', '}', '.', '!']
    return re.sub('([{}])'.format(''.join(markdown_escape_characters)), r'\\\1', text)


is_main_menu_button_active = {}







#13
from handlers import handle_additional_info, handle_szoreg_info, handle_schools_info, handle_survey_chart
#from create_chart import handle_survey_chart

#@dp.message_handler(Command('text'))
dp.register_callback_query_handler(handle_additional_info, lambda query: json.loads(query.data)["type"] == "additional_info")
#dp.register_callback_query_handler(handle_espd_info, lambda query: json.loads(query.data)["type"] == "espd_info")
dp.register_callback_query_handler(handle_szoreg_info, lambda query: json.loads(query.data)["type"] == "szoreg_info")
dp.register_callback_query_handler(handle_schools_info, lambda query: json.loads(query.data)["type"] == "schools_info")
#dp.register_callback_query_handler(handle_digital_ministry_info, lambda query: json.loads(query.data)["type"] == "digital_ministry_info")
#dp.register_callback_query_handler(handle_digital_ministry_info_post, lambda query: json.loads(query.data)["type"] == "digital_ministry_info_post")
dp.register_callback_query_handler(handle_survey_chart, lambda query: json.loads(query.data)["type"] == "survey_chart")

















import difflib


import asyncio


#@dp.message_handler(Command('text'))


import random
from aiogram import types




user_messages = {}


COLUMNS_TO_EXPORT = [1, 2, 3, 6, 7]

TABLE_HEADERS = ["Наименование", "Население", "Сотовая связь", "Интернет", "Таксофон"]


async def handler_otpusk_message(message, employees_on_vacation):
    if len(employees_on_vacation) > 0:
        response = "Сотрудники, которые сегодня в отпуске:\n\n"
        for employee in employees_on_vacation:
            response += f"{employee[0]} ({employee[1]})\n"
        time.sleep(2)
        await message.reply(response)
    else:
        time.sleep(2)
        await message.reply("Сегодня никто не в отпуске.")









from aiogram import executor
async def on_startup(dp):
    #await clear_cache()
    print("Бот запущен")

async def clear_cache():
    # Добавьте asyncio.loop.run_in_executor для работы с синхронным кодом в асинхронной функции
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, storage2.flushdb)
    print("Cache cleared")


if __name__ == "__main__":
    # Настройка и запуск планировщика
    scheduler = AsyncIOScheduler()
    scheduler.add_job(clear_cache, 'cron', hour=0, minute=0)
    scheduler.start()
    from handlers import *
    executor.start_polling(dp, on_startup=on_startup)



